import argparse
import os
import requests
import exiftool
import io
import zipfile
import pathlib
import xml.etree.ElementTree as ET
from glob import glob
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from threading import Thread
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich.progress import Progress, SpinnerColumn, TextColumn, DownloadColumn, BarColumn
from utils.colors import Color
from utils.folder import Folder
from crawler.data import DBData, Pages, Metadata, Files
import utils.ui as ui


class Reader:


    def __init__(self, seed:str):
        parsed_url = urlparse(seed)
        domain = parsed_url.netloc.replace(f":{parsed_url.port}","")
        self._output = os.path.join(Folder.get_local_folder(), domain)
        self._files_folder = os.path.join(self._output,'files')
        self._WORDS = ["user","password","domain","host","login","usuario","contraseña","dominio","code","page","print","impresora"]
    
    def _read_word(self, zip_file, document_name)->str:
        text=""
        with zip_file.open(document_name) as xml:
            
            tree = ET.parse(xml)
            root = tree.getroot()
            for elem in root.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t"):
                if elem.text:                                                                                                                                                                                                              
                    text += elem.text + "\n"
        return text
    
    def read_pdf(self, file_name:str):
        text = ""
        with open(file_name, 'rb') as file:
            pdf = io.BytesIO(file.read())
            reader = PdfReader(pdf)
            total_pages = len(reader.pages)
            pages = min(total_pages, 10)
            for page in reader.pages:
                text+=page.extract_text()+'\n'
        
        return text

    def read_word(self,file_name):
        with open(file_name, 'rb') as file:
            with zipfile.ZipFile(io.BytesIO(file.read())) as zip_file:
                 if 'word/document.xml' in zip_file.namelist():
                    text = self._read_word(zip_file, 'word/document.xml')
        return text

    def read_excel(self, file_name:str):
        with open(file_name, 'rb') as file:
            xlsx = io.BytesIO(file.read())
            wb = load_workbook(xlsx)
            ws = wb.active
            text = ''
            
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        text += str(cell) + ','
                
                if len([v for v in row if v is not None])>0:
                    text+='\n'
       
        return text


    def read(self):
        console = Console()
        for file in glob(os.path.join(self._files_folder,"*.*")):
            text = ""
            ext = pathlib.Path(file).suffix
            if ext == '.pdf':
                text = self.read_pdf(file)
            elif ext == '.docx':
                text = self.read_word(file)
            elif ext == '.xlsx':
                text = self.read_excel(file)
            for word in  self._WORDS:
                
                text_color,formatted, positions = Color.highlight_text(text, word, 'yellow')
                if formatted:
                    
                    #console.print(file, style="black on white")
                    message = ""
                    for p in positions:
                        message+=f"[green][*][reset] {(text_color[p[0]-30:p[1]+100]).replace('\n',' ')}\r\n"
                    
                    console.print(Panel(message, title=file))
                        #Color.print(f"[green][*][reset] {text_color[p[0]-20:p[1]+100]}")
            

class Crawler:
    def __init__(self, seed:str):
        if not seed.startswith("http"):
            seed = f"http://{seed}" 
        self._seed = seed
        self._visited = set()
        parsed_url = urlparse(seed)
        domain = parsed_url.netloc.replace(f":{parsed_url.port}","")
        self._output = os.path.join(Folder.get_local_folder(), domain)
        self._file_export = None
        self._files_folder = os.path.join(self._output,'files')
        self._summary = []
        self._results = []
        self._EXTENSIONS = [".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",".json",".txt"]
        self._ALLOWED_PAGES = [".aspx",".html",".htm",".jsp",".php",".php5",".asp"]
        self._table_name = "Pages"
        self._db = None
        self._executor = ThreadPoolExecutor(max_workers=1)
        self._progress = Progress(SpinnerColumn(),
                        BarColumn(),
                        TextColumn("[bold blue]{task.description}:[/] {task.fields[url]}"),
                        DownloadColumn())
        self._task = self._progress.add_task("[green]Indexing...", url="...", total=None)
        self._downloaded_task = self._progress.add_task("[yellow]Downloading...", url="...", total=100, visible=False)
        self._threads = []
        self._console = Console()
        self._init()
    
    def _init(self):
        Folder.create_folder(self._output)
        Folder.create_folder(self._files_folder)
        db_file = os.path.join(self._output,'data.db')
        self._db = DBData(db_file)
        
    def _is_valid(self, url:str, extensions)->bool:
        return any(url.lower().endswith(ext) for ext in extensions)

    def _is_document_valid(self, url) -> bool:
        return self._is_valid(url, self._EXTENSIONS)
    
    def _is_url_valid(self, url:str) -> bool:
        isvalid = self._is_valid(url, self._ALLOWED_PAGES)
        if not isvalid:
            path = urlparse(url).path
            _,ext = os.path.splitext(path)
            isvalid = ext == ""
        
        return isvalid

    def _print_users(self,  users):
        ui.print_table(users, exclude_columns=['id','key'], key_column="value")
        authors = set()
        users_output = None
        if self._file_export:
            users_output = open(self._file_export, '+a')
            
        for item in users:
            if item.value not in authors and item.value:
                authors.add(item.value)
                if users_output:
                    users_output.write(f'{item.value}\n')

    
    def _get_metadata(self, file_name:str) -> dict:
        
        metadata = None
        with exiftool.ExifTool() as et:
            metadata = et.execute_json(file_name)
            author = None
            creator = None
            producer = None

            for properties in metadata:
                if "PDF:Author" in properties:
                    author = properties["PDF:Author"]
                if "PDF:Creator" in properties:
                    creator = properties["PDF:Creator"]
                if "PDF:Producer" in properties:
                    producer = properties["PDF:Producer"]
                
                information = {
                    "author":author,
                    "producer":producer,
                    "creator":creator,
                    "filename":file_name
                }

                self._results.append(information)
        
        return metadata

    def _download_file(self, url:str):
        query = self._db.select_files(url)
        if len(query) == 0:
            file_name = os.path.basename(urlparse(url).path)
            self._progress.update(self._downloaded_task, url=file_name, advance=0,total=100, visible=True)
            target_path = os.path.join(self._files_folder,file_name)
            if not os.path.exists(target_path):
                resp = requests.get(url, stream=True)
                resp.raise_for_status()
                self._progress.update(self._downloaded_task, total=int(resp.headers.get('Content-Length', 0)))
                with open(target_path, "wb") as f:
                    for chunk in resp.iter_content(chunk_size=8192):
                        f.write(chunk)
                        self._progress.update(self._downloaded_task, advance=len(chunk))

                # TODO: Change
            information = self._get_metadata(target_path)
            file_id = self._db.add(Files(url=url))
            
            list = []
            for info in information:
                for property in info:
                    list.append(Metadata(file_id=file_id, key=property, value=str(info[property])))
                
                self._db.add_all(list)
                self._progress.update(self._downloaded_task, visible=False)

    def start(self, url:str=None):

        if url is None:
            url = self._seed
        else:
            url = url.lower()
            
        if url in self._visited:
            return
        
        if not self._is_url_valid(url):
            return

        
        self._visited.add(url)
        page = self._db.select_page(url)
        if len(page) == 0:
            self._db.add(Pages(url=url))
        #Color.print_info(f'{url}', end='\r')
        #self._progress.print(f"[white][▼][bold]{url}[reset]")
        self._progress.update(self._task, url=url)
        abs_url = None
        try:
            self._summary.append(url)
            r = requests.get(url)
            r.raise_for_status()
            soup = BeautifulSoup(r.content, 'html.parser')
            for link in soup.find_all('a', href=True):
                href = link["href"]
                abs_url = urljoin(url, href)
                if self._is_document_valid(abs_url):
                    #self._executor.submit(self._download_file, abs_url)
                    self._download_file(abs_url)
                elif abs_url.startswith(self._seed):
                    self.start(abs_url)
                    
        except requests.exceptions.RequestException as e:
            self._progress.print(f"[X] {abs_url}")
            #self._progress.print(e)
        except Exception as e:
            self._progress.print(e)
    
    def comments(self):
        comments = self._db.get_metadata(["Comments"])
        ui.print_table(comments, exclude_columns=['id','key','file_id'])

    def users(self):
        users = self._db.get_users()
        if len(users)>0:
            self._print_users(users)
        else:
            self._console.print(f"[green][*][yellow] Users not found for [bold]{self._seed}")
    
    def url(self):
        files = self._db.get_files()
        if self._file_export:
            url_output = open(self._file_export, '+a')
            for file in files:
                url_output.write(file.url+"\n")

        ui.print_table(files)
    
    def all(self):
        row_users = []
        row_sw = []
        distinct_users = set()
        distinct_sw = set()
        users = self._db.get_users()
        comments = self._db.get_metadata(["Comments"])
        software = self._db.get_software()
        comments_length = len(comments)
        
        
        if len(users) >0:
            for u in users:
                if u.value not in distinct_users:
                    distinct_users.add(u.value)
                    row_users.append(u)

            
        
        if len(software)>0:
            for sw in software:
                if sw.value not in distinct_sw:
                    distinct_sw.add(sw.value)
                    row_sw.append(sw)
            
        self._console.print(f"[green][*] Users [bold]{len(row_users)}[reset]")
        self._console.print(f"[green][*] Software [bold]{len(row_sw)}[reset]")
        self._console.print(f"[green][*] Comments [bold]{comments_length}[reset]")


    def crawler(self, args):
        try:
            if args.output:
                self._file_export = args.output
            if args.enumerate == 'users':
                self.users()
            elif args.enumerate == 'comments':
                self.comments()
            elif args.enumerate == 'url':
                self.url()
            elif args.enumerate == 'all':
                self.all()
            elif args.read:
                reader = Reader(self._seed)
                reader.read()
            
            else:
                with self._progress:
                    self.start()
        except Exception as e:
            Color.print_error(e)
        finally:
            self._executor.shutdown(wait=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("-i","--input", required=True, help="Single domain or file with list of URls")
    parser.add_argument("-e","--enumerate",  required=False, choices=["users","software","comments","all","url"])
    parser.add_argument("-r","--read",  required=False, action='store_true')
    parser.add_argument("-o","--output",  required=False, type=str)
    args = parser.parse_args()
    domains = set()
    if os.path.isfile(args.input):
        with open(args.input, 'r') as input:
            for line in input:
                if line not in domains:
                    domains.add(line.strip())
    else:
        domains.add(args.input)
    
    if args.output:
        if os.path.exists(args.output):
            os.remove(args.output)
    

    executor = ThreadPoolExecutor(max_workers=len(domains))

    for domain in domains:
        ui.console.print(f"[yellow]{domain}[reset]")
        crawler = Crawler(domain)
        if args.read is None and args.enumerate is None:
            executor.submit(crawler.crawler, args)
        else:
            crawler.crawler(args)

    if args.read is None and args.enumerate is None:
        executor.shutdown(wait=True)
        #crawler.crawler()

if __name__ == '__main__':
    main()


